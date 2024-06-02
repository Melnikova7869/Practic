import PyPDF2
import openpyxl

def extract_sections_text(pdf_path):
    try:
        with open(pdf_path, "rb") as file:
            reader = PyPDF2.PdfReader(file)
            
            sections_text = {"7.1": "", "7.2": ""}
            current_section = None
            
            for page in reader.pages:
                text = page.extract_text()
                
                if "7.1 Основная литература" in text:
                    current_section = "7.1"
                    sections_text[current_section] += text.split("7.1 Основная литература", 1)[-1]
                elif "7.2 Дополнительная литература" in text:
                    current_section = "7.2"
                    sections_text[current_section] += text.split("7.2 Дополнительная литература", 1)[-1]
                    break 
            return sections_text["7.1"].strip(), sections_text["7.2"].strip()
    except FileNotFoundError:
        print("Файл не найден.")
        return "", ""

def write_text_to_excel(text1, text2, excel_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    
    row = 1
    ws.cell(row=row, column=1, value="Основная литература")
    for line in text1.split('\n'):
        row += 1
        ws.cell(row=row, column=1, value=line)

    row += 2 
    ws.cell(row=row, column=1, value="Дополнительная литература")
    for line in text2.split('\n'):
        row += 1
        ws.cell(row=row, column=1, value=line)
    
    wb.save(excel_path)
    print(f"Данные из пунктов 7.1 и 7.2 файла {pdf_path} были успешно записаны в {excel_path}")

pdf_path = "/Users/frolovs/Desktop/pdf/B2.1.1_Uchebnaya(proektnaya)_praktika.pdf"
excel_path = "/Users/frolovs/Desktop/excel/Shablon_razmetki (1).xlsx"

text1, text2 = extract_sections_text(pdf_path)
if text1 or text2:
    write_text_to_excel(text1, text2, excel_path)
else:
    print("Не удалось извлечь данные из PDF.")